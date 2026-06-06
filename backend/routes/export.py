"""
Export routes for generating Excel, PDF reports
"""
from flask import Blueprint, request, jsonify, send_file
from flask_login import login_required, current_user
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from io import BytesIO
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib import colors
import os
from ..database import db
from ..models import Client, Payment, Expense

export_bp = Blueprint('export', __name__, url_prefix='/api/export')


def _is_authorized_for_client(client_id):
    """Check if current user can access client data"""
    client = Client.query.get(client_id)
    if not client:
        return False
    return client.trainer_id == current_user.id or getattr(current_user, 'is_admin', False)


@export_bp.route('/backup-json', methods=['GET'])
@login_required
def backup_json():
    """Export account data as JSON for backup."""
    clients_query = Client.query
    payments_query = Payment.query
    expenses_query = Expense.query

    if not getattr(current_user, 'is_admin', False):
        clients_query = clients_query.filter_by(trainer_id=current_user.id)
        payments_query = payments_query.filter_by(trainer_id=current_user.id)
        expenses_query = expenses_query.filter_by(trainer_id=current_user.id)

    payload = {
        'generated_at': datetime.utcnow().isoformat(),
        'clients': [
            {
                'id': c.id,
                'trainer_id': c.trainer_id,
                'name': c.name,
                'contact_number': c.contact_number,
                'email': c.email,
                'status': c.status,
                'pt_tier': c.pt_tier,
                'time_slot': c.time_slot,
                'renewal_date': c.renewal_date.isoformat() if c.renewal_date else None,
                'notes': c.notes,
                'created_at': c.created_at.isoformat() if c.created_at else None,
            }
            for c in clients_query.order_by(Client.created_at.desc()).all()
        ],
        'payments': [
            {
                'id': p.id,
                'client_id': p.client_id,
                'trainer_id': p.trainer_id,
                'amount': float(p.amount),
                'plan_type': p.plan_type,
                'start_date': p.start_date.isoformat() if p.start_date else None,
                'payment_date': p.payment_date.isoformat() if p.payment_date else None,
                'description': p.description,
                'created_at': p.created_at.isoformat() if p.created_at else None,
            }
            for p in payments_query.order_by(Payment.created_at.desc()).all()
        ],
        'expenses': [
            {
                'id': e.id,
                'trainer_id': e.trainer_id,
                'expense_name': e.expense_name,
                'category': e.category,
                'amount': float(e.amount),
                'expense_date': e.expense_date.isoformat() if e.expense_date else None,
                'created_at': e.created_at.isoformat() if e.created_at else None,
            }
            for e in expenses_query.order_by(Expense.created_at.desc()).all()
        ],
    }

    return jsonify(payload), 200


@export_bp.route('/clients-excel', methods=['GET'])
@login_required
def export_clients_excel():
    """Export all clients to Excel"""
    try:
        query = Client.query
        if not getattr(current_user, 'is_admin', False):
            query = query.filter_by(trainer_id=current_user.id)
        
        clients = query.all()
        
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Clients"
        
        # Header row
        headers = ['ID', 'Name', 'Status', 'PT Tier', 'Contact', 'Email', 'Renewal Date', 'Expected Amount', 'Created Date']
        ws.append(headers)
        
        # Style header
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
        
        # Add data rows
        for client in clients:
            ws.append([
                client.id,
                client.name,
                client.status,
                client.pt_tier,
                client.contact_number or '',
                client.email or '',
                client.renewal_date.isoformat() if client.renewal_date else '',
                client.expected_amount,
                client.created_at.strftime('%Y-%m-%d') if client.created_at else ''
            ])
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 5
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 10
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 20
        ws.column_dimensions['G'].width = 15
        ws.column_dimensions['H'].width = 15
        ws.column_dimensions['I'].width = 15
        
        # Save to bytes
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f"clients_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        )
    except Exception as e:
        return jsonify({'error': str(e)}), 400


@export_bp.route('/payments-excel', methods=['GET'])
@login_required
def export_payments_excel():
    """Export payments to Excel"""
    try:
        query = Payment.query
        if not getattr(current_user, 'is_admin', False):
            query = query.filter_by(trainer_id=current_user.id)
        
        payments = query.order_by(Payment.payment_date.desc()).all()
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Payments"
        
        headers = ['ID', 'Client', 'Amount', 'Payment Date', 'Description', 'Plan Type', 'Created Date']
        ws.append(headers)
        
        header_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
        
        for payment in payments:
            ws.append([
                payment.id,
                payment.client.name,
                float(payment.amount),
                payment.payment_date.isoformat(),
                payment.description or '',
                payment.plan_type,
                payment.created_at.strftime('%Y-%m-%d') if payment.created_at else ''
            ])
        
        ws.column_dimensions['A'].width = 5
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 20
        ws.column_dimensions['F'].width = 15
        ws.column_dimensions['G'].width = 15
        
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f"payments_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        )
    except Exception as e:
        return jsonify({'error': str(e)}), 400


@export_bp.route('/expenses-excel', methods=['GET'])
@login_required
def export_expenses_excel():
    """Export expenses to Excel"""
    try:
        query = Expense.query
        if not getattr(current_user, 'is_admin', False):
            query = query.filter_by(trainer_id=current_user.id)
        
        expenses = query.order_by(Expense.expense_date.desc()).all()
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Expenses"
        
        headers = ['ID', 'Expense Name', 'Category', 'Amount', 'Date', 'Created Date']
        ws.append(headers)
        
        header_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
        
        for expense in expenses:
            ws.append([
                expense.id,
                expense.expense_name,
                expense.category,
                float(expense.amount),
                expense.expense_date.isoformat(),
                expense.created_at.strftime('%Y-%m-%d') if expense.created_at else ''
            ])
        
        ws.column_dimensions['A'].width = 5
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 15
        
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f"expenses_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        )
    except Exception as e:
        return jsonify({'error': str(e)}), 400


@export_bp.route('/client-report/<int:client_id>', methods=['GET'])
@login_required
def export_client_report(client_id):
    """Export comprehensive client report (PDF)"""
    try:
        if not _is_authorized_for_client(client_id):
            return jsonify({'error': 'Unauthorized'}), 403
        
        client = Client.query.get(client_id)
        payments = Payment.query.filter_by(client_id=client_id).order_by(Payment.payment_date.desc()).all()
        
        # Create PDF
        pdf_buffer = BytesIO()
        doc = SimpleDocTemplate(pdf_buffer, pagesize=letter)
        story = []
        
        styles = getSampleStyleSheet()
        
        # Title
        title = Paragraph(f"<b>Client Report: {client.name}</b>", styles['Title'])
        story.append(title)
        story.append(Spacer(1, 0.3))
        
        # Client Info
        client_info = [
            ['Status:', client.status],
            ['PT Tier:', client.pt_tier],
            ['Email:', client.email or 'N/A'],
            ['Contact:', client.contact_number or 'N/A'],
            ['Renewal Date:', client.renewal_date.isoformat() if client.renewal_date else 'N/A'],
            ['Expected Monthly Amount:', f"₹{client.expected_amount}"],
        ]
        
        info_table = Table(client_info)
        info_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (0, -1), colors.lightgrey),
            ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        story.append(info_table)
        story.append(Spacer(1, 0.5))
        
        # Payments table
        story.append(Paragraph("<b>Payment History</b>", styles['Heading2']))
        
        payment_data = [['Date', 'Amount', 'Description', 'Plan Type']]
        for payment in payments[-10:]:  # Last 10 payments
            payment_data.append([
                payment.payment_date.isoformat(),
                f"₹{payment.amount}",
                payment.description or '',
                payment.plan_type
            ])
        
        payment_table = Table(payment_data)
        payment_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.blue),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey])
        ]))
        
        story.append(payment_table)
        
        # Generate PDF
        doc.build(story)
        pdf_buffer.seek(0)
        
        return send_file(
            pdf_buffer,
            mimetype='application/pdf',
            as_attachment=True,
            download_name=f"client_report_{client.name.replace(' ', '_')}_{datetime.now().strftime('%Y%m%d')}.pdf"
        )
    except Exception as e:
        return jsonify({'error': str(e)}), 400
